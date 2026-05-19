#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Tekla -> Macro Allied
=====================
Procesa archivos Excel exportados desde Tekla Structures
y los integra en la macro oficial de Allied.

Pensado para correr desde el nodo "Execute Command" de N8N:
    python tekla_to_allied.py

- Logs informativos -> stderr (visibles en N8N como log de ejecucion)
- Resultado final (JSON) -> stdout (lo parsea el siguiente nodo)
- Codigo de salida 0 si todo bien, 1 si hubo error fatal

Configuracion: editar las constantes al inicio o usar variables de entorno
    TEKLA_FOLDER, MACRO_PATH, OUTPUT_FOLDER
"""

import openpyxl
import os
import sys
import json
import shutil
import traceback
from datetime import datetime

# ── CONFIGURACION ─────────────────────────────────────────────
# Las rutas se toman de variables de entorno definidas en docker-compose.yml.
# Los valores por defecto son las rutas DENTRO del contenedor.
CARPETA_TEKLA      = os.environ.get("TEKLA_FOLDER",  "/data/tekla")
RUTA_MACRO         = os.environ.get("MACRO_PATH",    "/data/macro/Allied_Macro.xlsx")
CARPETA_SALIDA     = os.environ.get("OUTPUT_FOLDER", "/data/output")
CARPETA_PROCESADOS = os.path.join(CARPETA_TEKLA, "procesados")

# ── MAPEOS ────────────────────────────────────────────────────
TAB_MAP = {
    "SBS_Eave_Struts_Shipper":        "Eave Struts",
    "SBS_CEE_Secondary_Shipper":      "Cold Form Members (CEE)",
    "SBS_ZEE_Secondary_Shipper":      "Cold Form Members (ZEE)",
    "SBS_Miscellaneous_Shipper":      "Misc. Cold Form",
    "SBS_Clips_Shipper":              "Clips",
    "SBS_Pre_Galv_Clips_Shipper":     "Pre-Galv Clips",
    "Standing_Seam_Hardware_Shipper": "Standing Seam Hardware",
    "SBS_Screws_Shipper":             "Screws",
}
TAB_ZEE_EXTRA = ["Cold Form Members (ZEE) (2)", "Cold Form Members (ZEE) (3)"]

DESC_MAP = {
    "EAVE_STRUT":                      "Eave Strut (LSSS)",
    "SHEETING_BASE_CHANNEL":           "Base Angle",
    "SHEETING_ANGLE":                  "Sheeting Angle",
    "SHEETING_BASE_ANGLE":             "Base Angle",
    "C_WRAP_CHANNEL":                  "Girt Header (8 1/4CX6X4)",
    "C_STRUT_SPACER_LOW":              "Eave Strut Spacer  ( 3 : 12)",
    "STRAPPING":                       "Rolls of Strapping",
    "CCF_CLIP":                        "Clip",
    "CCF_CL5":                         "Sheeting Clip",
    "CCF_CL103":                       "Girt / Jamb Clip",
    "CCF_CL104":                       "Jamb Base Clip",
    "CCF_CL100":                       "Header to Jamb Clip",
    "FRAMED OPENING JAMB / SUB JAMB":  "Jamb",
    "FRAME OPENING JAMB / SUB JAMB":   "Jamb",
    "WALL GIRT":                       "Wall Girt",
    "ROOF PURLIN":                     "Roof Purlin",
    "FRAME OPENING HEADER":            "Frame Opening Header",
}

PART_MAP = {
    "L1225E14":     "12.25E14",
    "L4X2X16GA":    "4X2X16Ga",
    "SSL10_(3)":    "10X35C14",
    "WRAP8X4X14GA": "8X25C16",
}

# ── UTILIDADES DE LOG ─────────────────────────────────────────
def log(msg):
    """Logs informativos -> stderr (no contaminan el JSON de stdout)."""
    print(msg, file=sys.stderr, flush=True)

def output_json(data):
    """Resultado final -> stdout para que N8N lo parsee."""
    print(json.dumps(data, default=str, ensure_ascii=False))

# ── TRANSFORMACIONES ──────────────────────────────────────────
def t_desc(v):
    if not v: return ""
    k = str(v).strip().upper()
    if k in DESC_MAP: return DESC_MAP[k]
    return str(v).strip().replace("_", " ").title()

def t_part(v):
    if not v: return ""
    k = str(v).strip().upper()
    if k in PART_MAP: return PART_MAP[k]
    if k.startswith("L") and len(k) > 1 and k[1].isdigit():
        return str(v).strip()[1:]
    return str(v).strip()

def t_color(v):
    if v is None or str(v).strip() in ("0", ""): return ""
    return str(v).strip()

def t_dwg(v):
    if v is None or str(v).strip() in ("0", ""): return ""
    return str(v).strip()

def t_length(v):
    if v is None: return v
    s = str(v).strip()
    # Tekla HTML appends a spurious trailing `" ` after the real closing quote
    # e.g. `20'-7 3/8" "` → `20'-7 3/8"`
    while len(s) >= 2 and s.endswith(' "'):
        s = s[:-2].rstrip()
    return s

def to_float(v):
    """Convierte a float si es posible (para columna WT.)."""
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return v

_DETALLE_PREFIXES = ("IFLG=", "WEB=", "IFLG =", "WEB =", "WEB :", "FLG :")

def agregar_pulgadas(texto):
    """Formatea una fila de detalle Tekla añadiendo " a los números.

    Soporta los formatos que Tekla exporta:
      'IFlg=3 10.75'   -> 'IFlg=  3"   10.75"'
      'Web=22.1875'    -> 'Web=  22.1875"'
      'Web : 1.75 16.125' -> 'Web :  1.75"   16.125"'
      'Flg : 71.9375'  -> 'Flg :  71.9375"'
    Si los valores ya tienen " se conservan tal cual.
    """
    if not texto: return texto
    texto = str(texto).strip()
    prefijo, resto = "", texto
    for p in ["IFlg=", "Web=", "IFlg =", "Web =", "Web :", "Flg :"]:
        if texto.upper().startswith(p.upper()):
            prefijo = texto[:len(p)]
            resto = texto[len(p):].strip()
            break
    partes = resto.split()
    resultado = []
    for parte in partes:
        valor = parte.rstrip('"')
        try:
            float(valor)
            resultado.append(f'{valor}"')
        except ValueError:
            resultado.append(parte)
    return prefijo + "  " + "   ".join(resultado)

def es_detalle(fila):
    for c in fila:
        if c and str(c).strip().upper().startswith(_DETALLE_PREFIXES):
            return True
    return False

def es_peso(fila):
    return any(c and "PAGE WEIGHT" in str(c).upper() for c in fila)

def extraer_peso(fila):
    encontrado = False
    for c in fila:
        if c and "PAGE WEIGHT" in str(c).upper():
            encontrado = True
            continue
        if encontrado and c is not None and str(c).strip() != "":
            try: return float(c)
            except: return c
    return None

def concat_detalle(fila):
    return " ".join(str(c).strip() for c in fila if c and str(c).strip())

# ── LECTURA DE ARCHIVO EXCEL (.xls binario, .xlsx o HTML) ─────
def cargar_archivo_excel(ruta):
    """
    Carga un archivo Excel y devuelve un workbook openpyxl.
    Detecta el formato REAL por los primeros bytes (no por extension):
    - HTML disfrazado de .xls (lo que Tekla exporta): convierte con pandas
    - .xls binario real (Excel 97-2003 OLE): convierte con xlrd
    - .xlsx (Office Open XML): abre directo con openpyxl
    """
    with open(ruta, 'rb') as f:
        head = f.read(64)
    head_stripped = head.lstrip().lower()

    # Caso 1: HTML disfrazado de Excel (lo que entrega Tekla)
    if (head_stripped.startswith(b'<html') or
        head_stripped.startswith(b'<!doctype') or
        head_stripped.startswith(b'<table') or
        b'<html' in head_stripped):
        return _html_a_workbook(ruta)

    # Caso 2: .xlsx real (firma ZIP)
    if head.startswith(b'PK\x03\x04'):
        return openpyxl.load_workbook(ruta, data_only=True)

    # Caso 3: .xls binario real (firma OLE Compound Document)
    if head.startswith(b'\xd0\xcf\x11\xe0'):
        return _xls_binario_a_workbook(ruta)

    raise ValueError(
        f"Formato no reconocido en {os.path.basename(ruta)}. "
        f"Primeros bytes: {head[:16]!r}"
    )


def _xls_binario_a_workbook(ruta):
    """Lee un .xls binario real (Excel 97-2003) usando xlrd."""
    import xlrd
    xls_book = xlrd.open_workbook(ruta)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name in xls_book.sheet_names():
        xls_sheet = xls_book.sheet_by_name(sheet_name)
        ws = wb.create_sheet(sheet_name)
        for row_idx in range(xls_sheet.nrows):
            for col_idx in range(xls_sheet.ncols):
                cell_type = xls_sheet.cell_type(row_idx, col_idx)
                value = xls_sheet.cell_value(row_idx, col_idx)
                if cell_type in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                    continue
                if cell_type == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate.xldate_as_datetime(value, xls_book.datemode)
                elif cell_type == xlrd.XL_CELL_BOOLEAN:
                    value = bool(value)
                elif cell_type == xlrd.XL_CELL_ERROR:
                    value = None
                ws.cell(row=row_idx + 1, column=col_idx + 1, value=value)
    return wb


def _html_a_workbook(ruta):
    """
    Lee un archivo HTML disfrazado de .xls y devuelve un workbook openpyxl.
    Es lo que Tekla exporta en su formato 'Excel': HTML con tablas adentro.
    """
    import pandas as pd

    # header=None: no asumir que la primera fila es encabezado;
    # keep_default_na=False y na_values=[]: no convertir strings a NaN
    tables = pd.read_html(ruta, header=None, keep_default_na=False, na_values=[])

    if not tables:
        raise ValueError(f"No se encontraron tablas HTML en {os.path.basename(ruta)}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # El nombre de la pestaña se deriva del nombre del archivo
    # (SBS_Eave_Struts_Shipper.xls -> SBS_Eave_Struts_Shipper)
    base_name = os.path.splitext(os.path.basename(ruta))[0]
    # openpyxl limita los nombres de hoja a 31 caracteres
    sheet_name = base_name[:31]
    ws = wb.create_sheet(sheet_name)

    # Tekla suele exportar una sola tabla; tomamos esa
    df = tables[0]

    for row_idx, row in enumerate(df.itertuples(index=False), start=1):
        for col_idx, value in enumerate(row, start=1):
            # Saltar valores vacios o NaN
            if value is None:
                continue
            if isinstance(value, float) and pd.isna(value):
                continue
            if isinstance(value, str) and value.strip() == "":
                continue
            ws.cell(row=row_idx, column=col_idx, value=value)
    return wb


def _set_cell_safe(ws, row, col, value):
    """
    Asigna un valor a una celda manejando rangos merged (celdas fusionadas).
    Si la celda objetivo es parte de un merge, escribe en la celda principal
    (top-left del rango). Si no, escribe directo.
    """
    cell = ws.cell(row, col)
    if isinstance(cell, openpyxl.cell.cell.MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                ws.cell(merged_range.min_row, merged_range.min_col).value = value
                return
        # Si la celda dice ser MergedCell pero no encontramos su rango, ignorar
        return
    cell.value = value


HEADER_ALIASES = {
    "DRAWING #": "DWG #",
    "DRAWING NO": "DWG #",
    "DRAWING NO.": "DWG #",
    "DWG": "DWG #",
    "WT": "WT.",
    "WEIGHT": "WT.",
}

FIELD_HEADERS = {
    "qty": "QTY",
    "mark": "MARK",
    "desc": "DESCRIPTION",
    "pitch": "PITCH",
    "part": "PART",
    "punch": "PUNCH",
    "dwg": "DWG #",
    "color": "COLOR",
    "length": "LENGTH",
    "wt": "WT.",
}


def normalizar_header(v):
    if v is None:
        return ""
    texto = " ".join(str(v).strip().upper().split())
    if texto.endswith(":"):
        texto = texto[:-1].strip()
    return HEADER_ALIASES.get(texto, texto)


def rango_merged(ws, row, col):
    coord = ws.cell(row, col).coordinate
    for merged_range in ws.merged_cells.ranges:
        if coord in merged_range:
            return merged_range
    return None


def buscar_fila_encabezados(ws):
    for fila in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 40)):
        headers = [normalizar_header(c.value) for c in fila if c.value]
        if "QTY" in headers:
            return fila[0].row
    return None


def mapa_columnas(ws, header_row=None):
    header_row = header_row or buscar_fila_encabezados(ws)
    if not header_row:
        return None, {}

    columnas = {}
    for cell in ws[header_row]:
        header = normalizar_header(cell.value)
        if header and header not in columnas:
            columnas[header] = cell.column
    return header_row, columnas


def valor_por_header(fila, columnas, header):
    col = columnas.get(FIELD_HEADERS.get(header, header))
    if not col or col > len(fila):
        return None
    return fila[col - 1]


def primer_detalle(fila):
    for c in fila:
        if not c:
            continue
        texto = str(c).strip()
        if texto.upper().startswith(_DETALLE_PREFIXES):
            return texto
    return None


def celda_valor_despues_de_rotulo(ws, cell):
    merged = rango_merged(ws, cell.row, cell.column)
    if merged:
        return merged.min_row, merged.max_col + 1
    return cell.row, cell.column + 1


def valor_despues_de_rotulo(ws, cell):
    etiqueta = normalizar_header(cell.value)
    for col in range(cell.column + 1, ws.max_column + 1):
        valor = ws.cell(cell.row, col).value
        if valor is None or str(valor).strip() == "":
            continue
        if normalizar_header(valor) == etiqueta:
            continue
        return valor
    return None


# ── LECTURA ENCABEZADO ────────────────────────────────────────
def leer_encabezado(ws):
    enc = {}
    for fila in ws.iter_rows(min_row=1, max_row=7):
        for c in fila:
            if not c.value: continue
            v = str(c.value).upper()
            valor = valor_despues_de_rotulo(ws, c)
            if valor is None: continue
            if "JOB NUMBER"      in v: enc["job"]      = valor
            if "ISSUE DATE"      in v: enc["fecha"]    = valor
            if "BUILDING NUMBER" in v: enc["edificio"] = valor
            if "BLDG DESCRIP"    in v: enc["descrip"]  = valor
            if "CUSTOMER"        in v: enc["cliente"]  = valor
    return enc

# ── LECTURA DE PIEZAS ─────────────────────────────────────────
def leer_piezas(ws, tipo):
    piezas, pieza, peso = [], None, None
    header_row, columnas = mapa_columnas(ws)
    if not header_row or "QTY" not in columnas:
        raise ValueError(f"No se encontro fila de encabezados QTY en '{ws.title}'")

    for fila in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if es_peso(fila):
            peso = extraer_peso(fila)
            break
        detalle = primer_detalle(fila)
        if detalle:
            if pieza:
                detalle_fmt = agregar_pulgadas(detalle)
                pieza.setdefault("detalles", []).append(detalle_fmt)
                pieza["detalle"] = "\n".join(pieza["detalles"])
            continue

        qty = valor_por_header(fila, columnas, "qty")
        mark_raw = valor_por_header(fila, columnas, "mark")
        has_qty = qty is not None and str(qty).strip() not in ("", "QTY")
        has_mark = mark_raw is not None and str(mark_raw).strip() not in ("", "MARK")
        if has_qty or has_mark:
            if pieza: piezas.append(pieza)
            color = t_color(valor_por_header(fila, columnas, "color"))
            if tipo in ("Clips", "Pre-Galv Clips") and not color:
                color = "Pre-Galvanized"
            pieza = {
                "detalle": None,
                "detalles": [],
                "qty": qty,
                "mark": mark_raw,
                "desc": t_desc(valor_por_header(fila, columnas, "desc")),
                "pitch": valor_por_header(fila, columnas, "pitch"),
                "part": t_part(valor_por_header(fila, columnas, "part")),
                "punch": valor_por_header(fila, columnas, "punch"),
                "color": color,
                "dwg": t_dwg(valor_por_header(fila, columnas, "dwg")),
                "length": t_length(valor_por_header(fila, columnas, "length")),
                "wt": to_float(valor_por_header(fila, columnas, "wt")),
            }

    if pieza: piezas.append(pieza)
    return piezas, peso

# ── ESCRITURA ─────────────────────────────────────────────────
def escribir_enc(mws, enc, num, total):
    for fila in mws.iter_rows(min_row=1, max_row=7):
        for c in fila:
            if not c.value: continue
            v = str(c.value).upper()
            row, col = celda_valor_despues_de_rotulo(mws, c)
            if "SHIPPER NUMBER" in v: _set_cell_safe(mws, row, col, num)
            if v.strip() == "OF":     _set_cell_safe(mws, row, col, total)
            if "JOB NUMBER"      in v and enc.get("job") is not None:      _set_cell_safe(mws, row, col, enc.get("job"))
            if "ISSUE DATE"      in v and enc.get("fecha") is not None:    _set_cell_safe(mws, row, col, enc.get("fecha"))
            if "BUILDING NUMBER" in v and enc.get("edificio") is not None: _set_cell_safe(mws, row, col, enc.get("edificio"))
            if "BLDG DESCRIP"    in v and enc.get("descrip") is not None:  _set_cell_safe(mws, row, col, enc.get("descrip"))
            if "CUSTOMER"        in v and enc.get("cliente") is not None:  _set_cell_safe(mws, row, col, enc.get("cliente"))

def escribir_piezas(mws, piezas, peso, tipo):
    header_row, columnas = mapa_columnas(mws)
    if not header_row or "QTY" not in columnas:
        raise ValueError(f"No se encontro fila de encabezados QTY en macro '{mws.title}'")

    # Desfusionar cualquier rango merged en el area de datos (fila 11+).
    # El template de la macro tiene merges placeholder ahi que rompen
    # la escritura. No son funcionales, solo visuales.
    rangos_a_desfusionar = []
    for merged_range in list(mws.merged_cells.ranges):
        if merged_range.min_row >= header_row + 2:
            rangos_a_desfusionar.append(str(merged_range))
    for rng_str in rangos_a_desfusionar:
        mws.unmerge_cells(rng_str)

    data_start = header_row + 2
    for row in mws.iter_rows(min_row=data_start, max_row=mws.max_row):
        for cell in row:
            cell.value = None

    r = data_start
    detail_col = min(columnas.values()) if columnas else 1
    for p in piezas:
        for field, header in FIELD_HEADERS.items():
            col = columnas.get(header)
            if col:
                mws.cell(r, col).value = p.get(field)
        r += 1

        detalles = p.get("detalles") or ([p["detalle"]] if p.get("detalle") else [])
        for detalle in detalles:
            mws.cell(r, detail_col).value = detalle
            r += 1

    if peso is not None:
        r += 1
        wt_col = columnas.get("WT.", 10)
        label_col = max(1, wt_col - 2)
        mws.cell(r, label_col).value = "PAGE WEIGHT:"
        mws.cell(r, wt_col).value = peso

# ── PROCESAMIENTO PRINCIPAL ───────────────────────────────────
def procesar():
    inicio = datetime.now()
    log_entries, errores, files_with_errors = [], [], []

    # 1) Validar carpeta y archivos
    if not os.path.isdir(CARPETA_TEKLA):
        return {
            "status": "error",
            "message": f"Carpeta de Tekla no existe: {CARPETA_TEKLA}",
            "duration_seconds": 0
        }

    archivos = sorted([f for f in os.listdir(CARPETA_TEKLA)
                       if f.lower().endswith((".xlsx", ".xls"))
                       and os.path.isfile(os.path.join(CARPETA_TEKLA, f))])

    if not archivos:
        return {
            "status": "no_files",
            "message": "No hay archivos para procesar",
            "duration_seconds": 0
        }

    log(f"Encontrados {len(archivos)} archivos para procesar")

    # 2) Validar macro
    if not os.path.isfile(RUTA_MACRO):
        return {
            "status": "error",
            "message": f"Macro Allied no existe: {RUTA_MACRO}",
            "duration_seconds": 0
        }

    # 3) Cargar macro (preservando VBA)
    try:
        macro_wb = openpyxl.load_workbook(RUTA_MACRO)
    except Exception as e:
        return {
            "status": "error",
            "message": f"No se pudo abrir la macro: {e}",
            "duration_seconds": (datetime.now()-inicio).seconds
        }

    zee_count = 0
    job_number = "SIN_NUMERO"
    total = len(archivos)
    procesados_ok = []  # archivos a mover a /procesados al final

    # 4) Procesar cada archivo
    for i, nombre in enumerate(archivos, 1):
        ruta_archivo = os.path.join(CARPETA_TEKLA, nombre)
        try:
            wb = cargar_archivo_excel(ruta_archivo)
            ws = wb.active
            tab_tekla = ws.title
            tab_macro = TAB_MAP.get(tab_tekla)

            # Manejo especial de multiples ZEE
            if tab_tekla == "SBS_ZEE_Secondary_Shipper":
                if zee_count > 0 and zee_count-1 < len(TAB_ZEE_EXTRA):
                    tab_macro = TAB_ZEE_EXTRA[zee_count-1]
                zee_count += 1

            if not tab_macro or tab_macro not in macro_wb.sheetnames:
                msg = f"SKIP [{nombre}] -> pestana '{tab_macro}' no encontrada"
                log_entries.append(msg); log(msg)
                continue

            enc = leer_encabezado(ws)
            if enc.get("job"):
                job_number = str(enc["job"]).split("_")[0].strip()

            piezas, peso = leer_piezas(ws, tab_macro)
            mws = macro_wb[tab_macro]
            escribir_enc(mws, enc, i, total)
            escribir_piezas(mws, piezas, peso, tab_macro)

            msg = f"OK [{nombre}] -> [{tab_macro}] | {len(piezas)} piezas | Peso: {peso}"
            log_entries.append(msg); log(msg)
            procesados_ok.append(nombre)

        except Exception as e:
            msg = f"ERROR [{nombre}]: {e}"
            log_entries.append(msg); errores.append(msg); log(msg)
            files_with_errors.append({"file": nombre, "error": str(e)})
            log(traceback.format_exc())

    # 5) Pestana COVER
    if "Cover" in macro_wb.sheetnames:
        cov = macro_wb["Cover"]
        for fila in cov.iter_rows():
            for c in fila:
                if not c.value: continue
                if "JOB NUMBER" in str(c.value).upper():
                    _set_cell_safe(cov, c.row, c.column+1, job_number)
                if "TOTAL NUMBER" in str(c.value).upper():
                    _set_cell_safe(cov, c.row, c.column+1, total)

    # 6) Pestana LOG
    if "LOG" in macro_wb.sheetnames:
        del macro_wb["LOG"]
    lws = macro_wb.create_sheet("LOG")
    lws["A1"] = "REPORTE DE PROCESAMIENTO - TEKLA -> MACRO ALLIED"
    lws["A2"] = f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    lws["A3"] = f"Archivos procesados: {total}"
    lws["A4"] = f"Errores: {len(errores)}"
    lws["A5"] = ""
    for idx, linea in enumerate(log_entries, 6):
        lws.cell(idx, 1).value = linea

    # 7) Guardar
    os.makedirs(CARPETA_SALIDA, exist_ok=True)
    salida = os.path.join(CARPETA_SALIDA, f"{job_number}_Secondary_Shipper.xlsx")
    macro_wb.save(salida)
    log(f"Guardado en: {salida}")

    # 8) Archivar archivos procesados (para no reprocesarlos en la siguiente corrida)
    os.makedirs(CARPETA_PROCESADOS, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    archivo_destino_ts = os.path.join(CARPETA_PROCESADOS, timestamp)
    os.makedirs(archivo_destino_ts, exist_ok=True)
    for nombre in procesados_ok:
        try:
            shutil.move(
                os.path.join(CARPETA_TEKLA, nombre),
                os.path.join(archivo_destino_ts, nombre)
            )
        except Exception as e:
            log(f"No se pudo mover {nombre}: {e}")

    duracion = (datetime.now() - inicio).seconds

    return {
        "status": "success" if not errores else "partial_success",
        "job_number": job_number,
        "files_processed": total,
        "files_with_errors": files_with_errors,
        "output_file": salida,
        "archive_folder": archivo_destino_ts,
        "duration_seconds": duracion,
        "log_entries": log_entries
    }


# ── ENTRY POINT ───────────────────────────────────────────────
if __name__ == "__main__":
    output_json({
        "status": "error",
        "message": (
            "La escritura directa .xlsx esta deshabilitada. "
            "Use scripts/run_xls_host.sh para generar el .xls final, "
            "o scripts/export_tekla_payload.py para generar solo el payload."
        )
    })
    sys.exit(1)
